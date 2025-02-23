import streamlit as st
import pandas as pd
from io import StringIO
import warnings
warnings.filterwarnings("ignore")

uploaded_file = st.file_uploader("Choose a file")
if uploaded_file is not None:

    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    wb = load_workbook(uploaded_file)
    for st_name in wb.sheetnames:
        st = wb[st_name]
        mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]

        for mcr in mcr_coord_list:
          if mcr[0]=='D':
            # print(mcr)
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = st.cell(row=min_row, column=min_col).value
            # print(top_left_cell_value)
            st.unmerge_cells(mcr)
            for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value

    wb.save('merged_tmp.xlsx')

    import pandas as pd
    dfs=pd.read_excel('merged_tmp.xlsx',sheet_name=None,skiprows=2)

    entries_to_remove = ('Balance', 'Sheet1')
    for k in entries_to_remove:
        dfs.pop(k, None)


    table=pd.concat(dfs)
    table.reset_index(drop=True,inplace=True)
    table.columns = ['Date', 'Day', 'Changeover', 'Shift', 'PT 1', 'PT 2', 'PT 3', 'PT 4', 'PT 5', 'PT 6']
    table=table.drop(columns=['Day', 'Changeover'])
    # table=table.dropna(how='all')

    table["Date"] = pd.date_range(start='01-01-2025',end='31-01-2026',freq='D')
    # table

    cols = ['PT 1', 'PT 2', 'PT 3', 'PT 4', 'PT 5', 'PT 6']

    employees=[]
    

    for x in pd.unique(table[cols].values.ravel('K')):
      if x == x: # false for nan values
          employees.append(x)

    employees.sort()
    # for x in employees:
    #    print(type(x))
    # import streamlit as st

    # st.text(str(pd.unique(table[cols].values.ravel('K'))))

    import streamlit as st

    options = st.multiselect(
        "Who's still here?",
        employees,
        employees,
    )
    absent = [x for x in employees if x not in options]
    st.text(str(absent))
    # df = df[~df['date'].isin(a)]
    table=table[~table.isin(absent)]

    # st.dataframe(table)

    table['Number on holiday'] = table[cols].notna().sum(1)

    table['Number on holiday'] = table[cols].notna().sum(1)

    def isSummer(x):
      if(x >= datetime(2025,6,1)) and x <= datetime(2025,8,31):
        return True
      else:
        return False

    from datetime import datetime,timedelta
    table['Summer'] = [isSummer(i) for i in table['Date']]

    def canHoliday(num,summer):
      if (summer == True and num < 4) or (summer == False and num <3):
        return True
      else:
        return False

    table['Holiday Available'] = [canHoliday(x, y) for x, y in zip(table['Number on holiday'], table['Summer'])]

    count=0
    counts=[]

    for index, row in table.iterrows():
      if row['Holiday Available']==True:
        count+=1
      else:
        counts.append([count,row['Date'],index])
        count=0



    # st.write("You selected:", options)



    curr=str("Last update: {}".format(datetime.now()))
    st.text(curr)

    # mnm=str("bleh")
    # st.text(mnm)
    # # st.dataframe(table)

    for item in counts:
      if item[0]>4:
        mnm=str("{} days available from {} to {}, {} shifts ({} hours) required".format(item[0],
        (item[1]-timedelta(days=item[0])).strftime("%B %d, %Y"),
          (item[1]-timedelta(days=1)).strftime("%B %d, %Y"),
            table.loc[item[2]-item[0]:item[2]-1,"Shift"].count(),
              (table.loc[item[2]-item[0]:item[2]-1,"Shift"].count())*12))
        st.text(mnm)